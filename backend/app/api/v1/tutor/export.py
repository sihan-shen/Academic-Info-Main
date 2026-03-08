"""
导师信息导出接口
支持导出为Excel和CSV格式（管理员权限）
"""

from fastapi import APIRouter, HTTPException, Depends, Request, Query, Response
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime
import io
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.models import User
from app.utils import (
    success_response,
    error_response,
    business_error_response,
    api_logger,
    get_admin_user
)
from app.db.mongo import get_db

router = APIRouter(
    prefix="/tutor",
    tags=["tutor_export"]
)


async def get_tutors_data(
    db,
    keyword: Optional[str] = None,
    school: Optional[str] = None,
    department: Optional[str] = None,
    title: Optional[str] = None,
    limit: int = 1000
):
    """
    获取导师数据用于导出
    
    Args:
        db: 数据库连接
        keyword: 搜索关键词
        school: 学校筛选
        department: 院系筛选
        title: 职称筛选
        limit: 最大导出数量
    
    Returns:
        导师数据列表
    """
    # 构建查询条件
    query = {
        "$or": [
            {"is_deleted": {"$exists": False}},
            {"is_deleted": False}
        ]
    }
    
    if keyword:
        query["$and"] = query.get("$and", [])
        query["$and"].append({
            "$or": [
                {"name": {"$regex": keyword, "$options": "i"}},
                {"research_direction": {"$regex": keyword, "$options": "i"}},
                {"school_name": {"$regex": keyword, "$options": "i"}},
                {"department_name": {"$regex": keyword, "$options": "i"}}
            ]
        })
    
    if school:
        query["school_name"] = {"$regex": school, "$options": "i"}
    
    if department:
        query["department_name"] = {"$regex": department, "$options": "i"}
    
    if title:
        query["title"] = {"$regex": title, "$options": "i"}
    
    # 查询导师数据
    tutors_cursor = db.tutors.find(query).limit(limit)
    tutors = await tutors_cursor.to_list(length=limit)
    
    # 转换为导出格式
    export_data = []
    for tutor in tutors:
        export_data.append({
            "ID": tutor.get("id", ""),
            "姓名": tutor.get("name", ""),
            "职称": tutor.get("title", ""),
            "学校": tutor.get("school_name", ""),
            "院系": tutor.get("department_name", ""),
            "研究方向": tutor.get("research_direction", ""),
            "邮箱": tutor.get("email", ""),
            "电话": tutor.get("phone", ""),
            "个人主页": tutor.get("personal_page_url", ""),
            "招生类型": {
                "academic": "学硕",
                "professional": "专硕",
                "both": "学硕+专硕"
            }.get(tutor.get("recruitment_type"), ""),
            "是否有经费": "是" if tutor.get("has_funding") else "否",
            "论文数量": tutor.get("paper_count", 0),
            "项目数量": tutor.get("project_count", 0),
            "标签": ", ".join(tutor.get("tags", [])),
            "创建时间": tutor.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if tutor.get("created_at") else "",
            "更新时间": tutor.get("updated_at").strftime("%Y-%m-%d %H:%M:%S") if tutor.get("updated_at") else ""
        })
    
    return export_data


def create_excel_file(data: List[dict]) -> io.BytesIO:
    """
    创建Excel文件
    
    Args:
        data: 导师数据列表
    
    Returns:
        Excel文件的字节流
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "导师信息"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    if data:
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_csv_file(data: List[dict]) -> io.StringIO:
    """
    创建CSV文件
    
    Args:
        data: 导师数据列表
    
    Returns:
        CSV文件的字符串流
    """
    output = io.StringIO()
    
    if data:
        headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return output


@router.get(
    "/admin/export",
    summary="导出导师信息（管理员）",
    description="将导师列表导出为Excel或CSV格式，支持筛选条件"
)
async def export_tutors(
    request: Request,
    format: str = Query("excel", regex="^(excel|csv)$", description="导出格式：excel 或 csv"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    school: Optional[str] = Query(None, description="学校筛选"),
    department: Optional[str] = Query(None, description="院系筛选"),
    title: Optional[str] = Query(None, description="职称筛选"),
    limit: int = Query(1000, ge=1, le=10000, description="最大导出数量"),
    admin_user: User = Depends(get_admin_user)
):
    """
    导出导师信息接口（管理员权限）
    
    支持：
    1. 导出为Excel格式（.xlsx）
    2. 导出为CSV格式（.csv）
    3. 支持筛选条件
    4. 最多导出10000条记录
    
    Args:
        request: 请求对象
        format: 导出格式（excel/csv）
        keyword: 搜索关键词
        school: 学校筛选
        department: 院系筛选
        title: 职称筛选
        limit: 最大导出数量
        admin_user: 当前管理员用户
    
    Returns:
        文件下载响应
    """
    try:
        db = get_db()
        
        # 获取导师数据
        tutors_data = await get_tutors_data(
            db=db,
            keyword=keyword,
            school=school,
            department=department,
            title=title,
            limit=limit
        )
        
        if not tutors_data:
            raise HTTPException(
                status_code=404,
                detail=business_error_response(
                    code="NO_DATA",
                    message="没有符合条件的导师数据"
                )
            )
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "excel":
            # 创建Excel文件
            excel_file = create_excel_file(tutors_data)
            filename = f"导师信息_{timestamp}.xlsx"
            
            api_logger.info(
                f"管理员 {admin_user.id} 导出导师信息（Excel）\n"
                f"导出数量: {len(tutors_data)}\n"
                f"筛选条件: keyword={keyword}, school={school}, department={department}, title={title}\n"
                f"Request ID: {request.state.request_id}"
            )
            
            return StreamingResponse(
                excel_file,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
        
        else:  # csv
            # 创建CSV文件
            csv_file = create_csv_file(tutors_data)
            filename = f"导师信息_{timestamp}.csv"
            
            # 转换为字节流并添加BOM（解决Excel打开CSV中文乱码问题）
            csv_content = csv_file.getvalue()
            csv_bytes = io.BytesIO()
            csv_bytes.write('\ufeff'.encode('utf-8'))  # 添加BOM
            csv_bytes.write(csv_content.encode('utf-8'))
            csv_bytes.seek(0)
            
            api_logger.info(
                f"管理员 {admin_user.id} 导出导师信息（CSV）\n"
                f"导出数量: {len(tutors_data)}\n"
                f"筛选条件: keyword={keyword}, school={school}, department={department}, title={title}\n"
                f"Request ID: {request.state.request_id}"
            )
            
            return StreamingResponse(
                csv_bytes,
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
        
    except HTTPException:
        raise
    except Exception as e:
        api_logger.error(
            f"导出导师信息失败: {str(e)}\n"
            f"管理员: {admin_user.id}\n"
            f"格式: {format}\n"
            f"Request ID: {request.state.request_id}"
        )
        raise HTTPException(
            status_code=500,
            detail=error_response(
                message="导出导师信息失败",
                error={"request_id": request.state.request_id}
            )
        )


@router.get(
    "/admin/export-stats",
    summary="获取可导出数据统计（管理员）",
    description="获取符合条件的导师数量，用于导出前预览"
)
async def get_export_stats(
    request: Request,
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    school: Optional[str] = Query(None, description="学校筛选"),
    department: Optional[str] = Query(None, description="院系筛选"),
    title: Optional[str] = Query(None, description="职称筛选"),
    admin_user: User = Depends(get_admin_user)
):
    """
    获取可导出数据统计接口（管理员权限）
    
    用于导出前预览符合条件的数据量
    
    Args:
        request: 请求对象
        keyword: 搜索关键词
        school: 学校筛选
        department: 院系筛选
        title: 职称筛选
        admin_user: 当前管理员用户
    
    Returns:
        数据统计信息
    """
    try:
        db = get_db()
        
        # 构建查询条件
        query = {
            "$or": [
                {"is_deleted": {"$exists": False}},
                {"is_deleted": False}
            ]
        }
        
        if keyword:
            query["$and"] = query.get("$and", [])
            query["$and"].append({
                "$or": [
                    {"name": {"$regex": keyword, "$options": "i"}},
                    {"research_direction": {"$regex": keyword, "$options": "i"}},
                    {"school_name": {"$regex": keyword, "$options": "i"}},
                    {"department_name": {"$regex": keyword, "$options": "i"}}
                ]
            })
        
        if school:
            query["school_name"] = {"$regex": school, "$options": "i"}
        
        if department:
            query["department_name"] = {"$regex": department, "$options": "i"}
        
        if title:
            query["title"] = {"$regex": title, "$options": "i"}
        
        # 统计数量
        total_count = await db.tutors.count_documents(query)
        
        # 统计各学校数量
        school_pipeline = [
            {"$match": query},
            {"$group": {"_id": "$school_name", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        school_stats_cursor = db.tutors.aggregate(school_pipeline)
        school_stats = await school_stats_cursor.to_list(length=10)
        
        # 统计各职称数量
        title_pipeline = [
            {"$match": query},
            {"$group": {"_id": "$title", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        title_stats_cursor = db.tutors.aggregate(title_pipeline)
        title_stats = await title_stats_cursor.to_list(length=20)
        
        api_logger.info(
            f"管理员 {admin_user.id} 查询导出统计\n"
            f"符合条件的导师数量: {total_count}\n"
            f"Request ID: {request.state.request_id}"
        )
        
        return success_response(
            data={
                "total_count": total_count,
                "max_export_limit": 10000,
                "can_export": total_count > 0,
                "school_stats": [
                    {"school": stat["_id"], "count": stat["count"]}
                    for stat in school_stats
                ],
                "title_stats": [
                    {"title": stat["_id"], "count": stat["count"]}
                    for stat in title_stats
                ]
            },
            message="获取导出统计成功"
        )
        
    except Exception as e:
        api_logger.error(
            f"获取导出统计失败: {str(e)}\n"
            f"管理员: {admin_user.id}\n"
            f"Request ID: {request.state.request_id}"
        )
        raise HTTPException(
            status_code=500,
            detail=error_response(
                message="获取导出统计失败",
                error={"request_id": request.state.request_id}
            )
        )
